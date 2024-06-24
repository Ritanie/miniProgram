import streamlit as st
import pandas as pd 
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime


# Sidebar options
selected_sidebar_option = st.sidebar.radio(
    "Choose one function to proceed",
    ("File Converter", "Weight Calculator")
)

def display_converter():
    st.title("txt file to Excel file Converter")
    st.subheader("Upload your txt file here")
    uploaded_file = st.file_uploader("Choose a txt file")
    if uploaded_file is not None:
        df = pd.read_csv(uploaded_file, delimiter='\t')
        st.write(df)

        df_sorted = df.sort_values(by='purchase-date', ascending=False)

        df_sorted = df_sorted.reset_index(drop=True)
        df_sorted.index = df_sorted.index + 1



        df1 = pd.DataFrame()


        # Create a dictionary to track the index for each buyer-name
        recipient_name_index = {}
        current_index = 1
        # List to store the new indices
        No = []
        for recipient_name in df_sorted['recipient-name']:
            if recipient_name not in recipient_name_index:
                recipient_name_index[recipient_name] = current_index
                current_index += 1
            No.append(recipient_name_index[recipient_name])
        #print(No)
        df1['No.'] = No



        Customer_Name = []
        for x in df_sorted['recipient-name']:
            Customer_Name.append(x)
        #print(Customer_Name)
        df1['Customer Name'] = Customer_Name



        SKU_No = []
        for x in df_sorted['sku']:
            SKU_No.append(x)
        #print(SKU_No)
        df1['SKU'] = SKU_No



        shape_dict = {
            'H02': '方形',
            'H03': '三角形',
            'H04': '竖版'
        }
        color_dict = {
            'BG': '米色',
            'BN': '棕色',
            'BL': '蓝色',
            'GR': '灰色',
            'SB': '沙色（黄色）'
        }
        def word_preprocess(sku: str) -> str:
            shape = ''
            color = ''
            # Check for shape code in the SKU
            for key in shape_dict:
                if key in sku:
                    shape = shape_dict[key]
                    break
            # Check for color code in the SKU
            for key in color_dict:
                if key in sku:
                    color = color_dict[key]
                    break
            
            # Combine the shape and color
            description = f"{shape}{color}"
            return description
        Description = [word_preprocess(sku) for sku in SKU_No]
        #print(Description)
        df1['Description'] = Description



        def extract_size(product_name: str) -> str:
            match = re.search(r"\d+' x \d+'(?: x \d+')?", product_name)
            if match:
                return match.group(0)
            else:
                return "Error"
        Product_Size = df_sorted['product-name'].apply(extract_size)
        #print(Product_Size)
        Size = []
        for x in Product_Size:
            Size.append(x)
        #print(Size)
        df1['Product Size'] = Size



        Quantity = []
        for x in df_sorted['quantity-purchased']:
            Quantity.append(x)
        #print(Quantity)
        df1['Quantity'] = Quantity





        st.write("This file has been converted successfully! Check it out below:")
        st.write(df1)
        st.write(f"""
            *Note: There are {max(No)} orders in total'.*
        """)
        st.subheader("Download File")
        #st.text_input("Please name your file:")
        name = str(datetime.today().strftime('%Y%m%d'))+'01'
        file_name = st.text_input(
            "Please name your file:",
            placeholder=name,
        )
        if file_name is '':
            file_name = name
        if st.button("Submit") and file_name is not None:
            st.write(f"This file will be downloaded as {file_name}.xlsx")

            #file_name = "2024062301"

            df1.to_excel(f"{file_name}.xlsx", index=False)


            wb = load_workbook(f"{file_name}.xlsx")
            ws = wb.active

            center_aligned_text = Alignment(horizontal="center", vertical="center")

            header_font = Font(name='Arial', size=16, bold=False)
            grey_background = PatternFill(fgColor="F3F3F3", fill_type="solid")


            No_Quant_font = Font(name='Arial', size=12, italic=False, bold=True)
            SKU_font = Font(name='Arial', size=12, italic=False, bold=False)
            Cust_Prod_font = Font(name='Arial', size=12, italic=False, color='006F83')
            Desc_font = Font(name='Aptos Narrow (Body)',  size=12, italic=False, bold=False)





            # Define the border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Calculate max width for each column including the header
            max_widths = [0] * ws.max_column

            # Apply styles to the header row
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = center_aligned_text
                cell.fill = grey_background
                cell.border = thin_border
                max_widths[cell.column - 1] = max(max_widths[cell.column - 1], len(str(cell.value)))

            # Apply styles to the content rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                ws.row_dimensions[row[0].row].height = 17
                for cell in row:
                    cell.alignment = center_aligned_text
                    cell.border = thin_border
                    if cell.column == 1 or cell.column == 6:  # Apply No_Quant_font to the first and last columns
                        cell.font = No_Quant_font
                    elif cell.column == 2 or cell.column == 5:  # Apply Cust_Prod_font to the Customer Name column
                        cell.font = Cust_Prod_font
                    elif cell.column == 3:  # Apply SKU_font to the SKU column
                        cell.font = SKU_font
                    elif cell.column == 4:  # Apply Desc_font to the Description column
                        cell.font = Desc_font
                    else:  # Apply a default font to other cells if necessary
                        cell.font = Font(name='Arial', size=12, italic=False, bold=False)

                    # Calculate the maximum width needed for the column
                    max_widths[cell.column - 1] = max(max_widths[cell.column - 1], len(str(cell.value)))

            ws.row_dimensions[1].height = 63
            # Adjust the column widths
            for i, width in enumerate(max_widths):
                ws.column_dimensions[chr(65 + i)].width = width + 10  # Adding some padding

            # Apply styles to the content rows, set row height, and calculate max width for each column
            previous_no_value = None
            start_row = 2

            for row in range(2, ws.max_row + 1):
                current_no_value = ws.cell(row=row, column=1).value

                # Check if the current "No." value is the same as the previous one
                if current_no_value == previous_no_value:
                    continue
                else:
                    # Merge cells in the "No." column if the previous and current values are different
                    if previous_no_value is not None and start_row < row - 1:
                        ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                        ws.cell(row=start_row, column=1).alignment = center_aligned_text
                        ws.cell(row=start_row, column=1).border = thin_border

                    # Update the start_row for the next group
                    start_row = row
                    previous_no_value = current_no_value

            # Ensure the last group of cells are merged
            if start_row < ws.max_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=ws.max_row, end_column=1)
                ws.cell(row=start_row, column=1).alignment = center_aligned_text
                ws.cell(row=start_row, column=1).border = thin_border



            # Save the changes
            wb.save(f"{file_name}.xlsx")

            # Read the styled Excel file as binary
            with open(f"{file_name}.xlsx", "rb") as file:
                file_data = file.read()

            # Download button for the styled Excel file
            st.download_button(
                label="Download file as Excel",
                data=file_data,
                file_name=f"{file_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
def display_calculator():
    st.title("Weight(lbs) Calculator")
    df2 = pd.read_excel("SizeWeight.xlsx")

    sizes = {}

    # Iterate over rows in the DataFrame to populate the dictionary
    for index, row in df2.iterrows():
        size_key = row['Size']
        sizes[size_key] = {
            "Package Size": row['Package Size'],
            "Weight(lb)": row['Weight(lb)']
        }
    s = []
    for key in sizes:
        s.append(key)

    history = {}


    #size_input ='12‘ x 19’'
    #shape = 'H03'
    #quantity = 1
    size = st.selectbox(
        "What is the size?",
        (s))

    shape = st.radio(
            "Select your shape option",
            options=["H02", "H03"],
        )
    quantity = st.slider("Select your input quantity:", 0, 100, 1)



    if st.button("Submit"):
        original_weight = 0
        package = ''
        for key in sizes:
            if key in size:
                if shape == 'H02':
                    original_weight = sizes[key]['Weight(lb)']*quantity
                elif shape == 'H03':
                    original_weight = sizes[key]['Weight(lb)']/2*quantity
        #print(original_weight)

        new_weight = 0
        if shape=='H03' or quantity>1:
            for i in df2['Weight(lb)']:
                if original_weight != i:
                    p = []
                    indl = []
                    for i,j in enumerate(df2['Weight(lb)']):
                        ps = j-float(original_weight)
                        if ps > 0:
                            p.append(ps)
                            indl.append(j)
                    for a,b in enumerate(p):
                        if b == min(p):
                            new_weight = indl[a]   
                else:
                    new_weight=original_weight
        else:
            new_weight=original_weight
        #print(new_weight)

        ind = 0
        for ind, val in enumerate(df2['Weight(lb)']):
            if float(new_weight) == float(val):
                package = df2['Package Size'].iloc[ind]
                break
        #print(package)
        #print(ind+2)


        history['1'] = {'package':package, 'weight':new_weight}
        st.write(f'Selected Size: {size}')
        st.write(f'Selected Shape: {shape}')
        st.write(f'Selected Quantity: {quantity}')
        st.subheader(f'Package Size: {package}')
        st.subheader(f'Total Weight: {new_weight} lbs')

        
        

if selected_sidebar_option == "File Converter":
    display_converter()
elif selected_sidebar_option == "Weight Calculator":
    display_calculator()
